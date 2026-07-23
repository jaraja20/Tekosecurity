"""TEKOSECURE reports API tests — Fase 5"""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://emergent-security-1.preview.emergentagent.com").rstrip("/")
SUPABASE_URL = "https://fsucygjqzskwtnynvgob.supabase.co"
SUPABASE_ANON_KEY = "sb_publishable_WcxQECMlstlvfuzPV7ag3Q_ynzG0hVh"
EMAIL = "ti@nasser.com.py"
PASSWORD = "NasserTi73491654"


@pytest.fixture(scope="module")
def token():
    r = requests.post(
        f"{SUPABASE_URL}/auth/v1/token",
        params={"grant_type": "password"},
        headers={"apikey": SUPABASE_ANON_KEY, "Content-Type": "application/json"},
        json={"email": EMAIL, "password": PASSWORD},
        timeout=15,
    )
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


# ---- /api/reports/summary ----
def test_summary_no_auth():
    r = requests.get(f"{BASE_URL}/api/reports/summary", timeout=15)
    assert r.status_code == 401


def test_summary_default_30d(token):
    r = requests.get(
        f"{BASE_URL}/api/reports/summary",
        headers={"Authorization": f"Bearer {token}"},
        params={"days": 30},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["success"] is True
    assert data["range"]["from"] is not None
    assert data["range"]["to"] is not None
    s = data["summary"]
    assert s["total"] > 0, "expected seed data (15 attacks)"

    # by_type is a list of {attack_type, count} sorted desc
    assert isinstance(s["by_type"], list) and len(s["by_type"]) > 0
    counts = [it["count"] for it in s["by_type"]]
    assert counts == sorted(counts, reverse=True)
    for it in s["by_type"]:
        assert "attack_type" in it and "count" in it

    # by_severity has all 4 keys in SEVERITY_ORDER order
    sevs = [it["severity"] for it in s["by_severity"]]
    assert sevs == ["CRITICAL", "HIGH", "MEDIUM", "LOW"]

    # top_source_ips
    assert isinstance(s["top_source_ips"], list)
    assert len(s["top_source_ips"]) <= 15

    # type_by_severity is a dict / matrix
    assert isinstance(s["type_by_severity"], dict)
    for t, sevmap in s["type_by_severity"].items():
        assert set(sevmap.keys()) >= {"CRITICAL", "HIGH", "MEDIUM", "LOW"}


def test_summary_empty_range(token):
    r = requests.get(
        f"{BASE_URL}/api/reports/summary",
        headers={"Authorization": f"Bearer {token}"},
        params={"date_from": "2020-01-01", "date_to": "2020-01-02"},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["success"] is True
    assert data["summary"]["total"] == 0


def test_summary_24h(token):
    r = requests.get(
        f"{BASE_URL}/api/reports/summary",
        headers={"Authorization": f"Bearer {token}"},
        params={"days": 1},
        timeout=30,
    )
    assert r.status_code == 200
    assert r.json()["success"] is True


# ---- /api/reports/export.xlsx ----
EXPECTED_SHEETS = ["Resumen", "Tipos por Severidad", "Top IPs", "Serie temporal", "Detalle"]


def _validate_xlsx(body: bytes):
    assert len(body) >= 8000, f"xlsx too small ({len(body)} bytes)"
    wb = load_workbook(io.BytesIO(body))
    assert wb.sheetnames == EXPECTED_SHEETS, f"sheets: {wb.sheetnames}"


def test_export_xlsx_with_header(token):
    r = requests.get(
        f"{BASE_URL}/api/reports/export.xlsx",
        headers={"Authorization": f"Bearer {token}"},
        params={"days": 30},
        timeout=60,
    )
    assert r.status_code == 200, r.text[:300]
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = r.headers.get("content-disposition", "")
    assert "attachment" in cd and "tekosecure_reporte_" in cd and ".xlsx" in cd
    _validate_xlsx(r.content)


def test_export_xlsx_with_token_qs(token):
    """<a download> path used by the frontend."""
    r = requests.get(
        f"{BASE_URL}/api/reports/export.xlsx",
        params={"days": 30, "token_qs": token},
        timeout=60,
    )
    assert r.status_code == 200, r.text[:300]
    _validate_xlsx(r.content)


def test_export_xlsx_no_token():
    r = requests.get(f"{BASE_URL}/api/reports/export.xlsx", params={"days": 30}, timeout=15)
    assert r.status_code == 401
