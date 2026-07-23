"""
TEKOSECURE — Reports module

Aggregates data from Supabase attacks table and generates:
  - JSON summaries for the dashboard reports page
  - XLSX exports (openpyxl) for offline analysis
"""

from __future__ import annotations

import io
import os
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY")


# ---------------------------------------------------------------------------
# Data access
# ---------------------------------------------------------------------------
async def fetch_attacks(
    user_token: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 2000,
) -> list[dict]:
    """
    Read attacks from Supabase filtered by created_at range.
    We paginate manually since the Supabase REST default cap is 1000.
    """
    params: dict[str, Any] = {
        "select": "*",
        "order": "created_at.desc",
    }
    if date_from:
        params["created_at"] = f"gte.{date_from}"
    if date_to:
        # PostgREST allows multiple filters on same column via key aliasing;
        # instead we chain with AND via `and=` for date_to
        params["and"] = f"(created_at.lte.{date_to})"

    rows: list[dict] = []
    page_size = min(limit, 1000)
    offset = 0
    async with httpx.AsyncClient(timeout=15) as client:
        while len(rows) < limit:
            r = await client.get(
                f"{SUPABASE_URL}/rest/v1/attacks",
                params=params,
                headers={
                    "apikey": SUPABASE_ANON_KEY,
                    "Authorization": f"Bearer {user_token}",
                    "Range-Unit": "items",
                    "Range": f"{offset}-{offset + page_size - 1}",
                },
            )
            if r.status_code >= 400:
                raise RuntimeError(f"supabase {r.status_code}: {r.text[:200]}")
            batch = r.json()
            rows.extend(batch)
            if len(batch) < page_size:
                break
            offset += page_size
    return rows[:limit]


# ---------------------------------------------------------------------------
# Summaries
# ---------------------------------------------------------------------------
SEVERITY_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]


def build_summary(attacks: list[dict]) -> dict:
    total = len(attacks)
    by_type: Counter[str] = Counter()
    by_severity: Counter[str] = Counter()
    by_status: Counter[str] = Counter()
    by_source_ip: Counter[str] = Counter()
    by_mikrotik: Counter[str] = Counter()
    per_day: dict[str, int] = defaultdict(int)
    # cross-tab: type × severity
    type_by_sev: dict[str, Counter[str]] = defaultdict(Counter)

    first_ts = None
    last_ts = None

    for a in attacks:
        t = (a.get("attack_type") or "UNKNOWN").upper()
        sev = (a.get("severity") or "LOW").upper()
        st = (a.get("status") or "UNKNOWN").upper()
        by_type[t] += 1
        by_severity[sev] += 1
        by_status[st] += 1
        type_by_sev[t][sev] += 1

        if a.get("source_ip"):
            by_source_ip[a["source_ip"]] += 1
        if a.get("mikrotik_ip"):
            by_mikrotik[a["mikrotik_ip"]] += 1

        ts = a.get("created_at")
        if ts:
            day = ts[:10]
            per_day[day] += 1
            if first_ts is None or ts < first_ts:
                first_ts = ts
            if last_ts is None or ts > last_ts:
                last_ts = ts

    return {
        "total": total,
        "range": {
            "first": first_ts,
            "last": last_ts,
        },
        "by_type": [
            {"attack_type": k, "count": v}
            for k, v in by_type.most_common()
        ],
        "by_severity": [
            {"severity": sev, "count": by_severity.get(sev, 0)}
            for sev in SEVERITY_ORDER
        ],
        "by_status": [
            {"status": k, "count": v}
            for k, v in by_status.most_common()
        ],
        "top_source_ips": [
            {"source_ip": k, "count": v}
            for k, v in by_source_ip.most_common(15)
        ],
        "by_mikrotik": [
            {"mikrotik_ip": k, "count": v}
            for k, v in by_mikrotik.most_common()
        ],
        "per_day": [
            {"date": d, "count": per_day[d]}
            for d in sorted(per_day)
        ],
        "type_by_severity": {
            t: {sev: type_by_sev[t].get(sev, 0) for sev in SEVERITY_ORDER}
            for t in by_type
        },
    }


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="0F1533")
_HEADER_FONT = Font(bold=True, color="00BFFF", name="Calibri")
_TITLE_FONT = Font(bold=True, size=14, color="00BFFF")
_SUBTITLE_FONT = Font(italic=True, size=10, color="8A94A6")
_SEV_FILL = {
    "CRITICAL": PatternFill("solid", fgColor="FFEBEE"),
    "HIGH": PatternFill("solid", fgColor="FFF3E0"),
    "MEDIUM": PatternFill("solid", fgColor="FFFDE7"),
    "LOW": PatternFill("solid", fgColor="E8F5E9"),
}


def _write_header(ws, row: int, cols: list[str]) -> None:
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)


def build_xlsx(
    attacks: list[dict],
    summary: dict,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> bytes:
    wb = Workbook()

    # ---- Sheet 1: Resumen ----
    ws = wb.active
    ws.title = "Resumen"

    ws.cell(row=1, column=1, value="TEKOSECURE — Reporte de Amenazas").font = _TITLE_FONT
    period = f"{date_from or '(inicio)'} → {date_to or '(hoy)'}"
    ws.cell(row=2, column=1, value=f"Periodo: {period}").font = _SUBTITLE_FONT
    ws.cell(
        row=3,
        column=1,
        value=f"Generado: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}",
    ).font = _SUBTITLE_FONT

    ws.cell(row=5, column=1, value="Total de amenazas registradas").font = _HEADER_FONT
    ws.cell(row=5, column=2, value=summary["total"])

    r = 7
    _write_header(ws, r, ["Tipo de amenaza", "Cantidad"])
    r += 1
    for row in summary["by_type"]:
        ws.cell(row=r, column=1, value=row["attack_type"])
        ws.cell(row=r, column=2, value=row["count"])
        r += 1

    r += 1
    _write_header(ws, r, ["Severidad", "Cantidad"])
    r += 1
    for row in summary["by_severity"]:
        c1 = ws.cell(row=r, column=1, value=row["severity"])
        ws.cell(row=r, column=2, value=row["count"])
        if row["severity"] in _SEV_FILL:
            c1.fill = _SEV_FILL[row["severity"]]
        r += 1

    r += 1
    _write_header(ws, r, ["Estado", "Cantidad"])
    r += 1
    for row in summary["by_status"]:
        ws.cell(row=r, column=1, value=row["status"])
        ws.cell(row=r, column=2, value=row["count"])
        r += 1

    _autosize(ws)

    # ---- Sheet 2: Tipos × Severidad ----
    ws2 = wb.create_sheet("Tipos por Severidad")
    _write_header(ws2, 1, ["Tipo"] + SEVERITY_ORDER + ["Total"])
    r = 2
    for t, sevs in summary["type_by_severity"].items():
        ws2.cell(row=r, column=1, value=t)
        total = 0
        for i, sev in enumerate(SEVERITY_ORDER, start=2):
            v = sevs.get(sev, 0)
            ws2.cell(row=r, column=i, value=v)
            total += v
        ws2.cell(row=r, column=len(SEVERITY_ORDER) + 2, value=total).font = Font(bold=True)
        r += 1
    _autosize(ws2)

    # ---- Sheet 3: Top IPs atacantes ----
    ws3 = wb.create_sheet("Top IPs")
    _write_header(ws3, 1, ["IP origen", "Cantidad de ataques"])
    r = 2
    for row in summary["top_source_ips"]:
        ws3.cell(row=r, column=1, value=row["source_ip"])
        ws3.cell(row=r, column=2, value=row["count"])
        r += 1
    _autosize(ws3)

    # ---- Sheet 4: Ataques por día ----
    ws4 = wb.create_sheet("Serie temporal")
    _write_header(ws4, 1, ["Fecha", "Cantidad"])
    r = 2
    for row in summary["per_day"]:
        ws4.cell(row=r, column=1, value=row["date"])
        ws4.cell(row=r, column=2, value=row["count"])
        r += 1
    _autosize(ws4)

    # ---- Sheet 5: Detalle completo ----
    ws5 = wb.create_sheet("Detalle")
    cols = [
        "id",
        "created_at",
        "attack_type",
        "severity",
        "status",
        "source_ip",
        "destination_ip",
        "mikrotik_ip",
        "details",
    ]
    _write_header(ws5, 1, cols)
    r = 2
    for a in attacks:
        for i, key in enumerate(cols, start=1):
            v = a.get(key)
            ws5.cell(row=r, column=i, value=v)
        # colour by severity
        sev = (a.get("severity") or "").upper()
        if sev in _SEV_FILL:
            for i in range(1, len(cols) + 1):
                ws5.cell(row=r, column=i).fill = _SEV_FILL[sev]
        r += 1
    _autosize(ws5)
    ws5.freeze_panes = "A2"

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def suggested_filename(date_from: Optional[str], date_to: Optional[str]) -> str:
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    if date_from and date_to:
        return f"tekosecure_reporte_{date_from[:10]}_a_{date_to[:10]}.xlsx"
    if date_from:
        return f"tekosecure_reporte_desde_{date_from[:10]}.xlsx"
    return f"tekosecure_reporte_{today}.xlsx"


def default_range(days: int = 30) -> tuple[str, str]:
    now = datetime.now(timezone.utc)
    start = now - timedelta(days=days)
    return start.isoformat(), now.isoformat()
